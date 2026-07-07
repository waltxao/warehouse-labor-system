import re
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from app.models import Warehouse, DailyRecord, WeeklyReport, UploadLog
from app.core.exceptions import AppException

WAREHOUSE_CODE_PATTERN = re.compile(r'^(\d{3,4}[A-Z]?|[A-Z]{2}\d{2})$')

METRIC_MAP = {
    "\u6bcf\u65e5\u7bc0\u7701\u4eba\u6578": "labor_savings",
    "\u6bcf\u65e5\u8282\u7701\u4eba\u6570": "labor_savings",
    "\u6bcf\u65e5\u7cfb\u7d71\u4eba\u6578": "system_headcount",
    "\u6bcf\u65e5\u7cfb\u7edf\u4eba\u6570": "system_headcount",
    "\u5be6\u969b\u51fa\u52e4\u4eba\u6578": "actual_attendance",
    "\u5b9e\u9645\u51fa\u52e4\u4eba\u6570": "actual_attendance",
    "\u5be6\u969b\u5de5\u4f5c\u9700\u6c42\u4eba\u6578": "required_headcount_so",
    "\u5b9e\u9645\u5de5\u4f5c\u9700\u6c42\u4eba\u6570": "required_headcount_so",
    "\u5be6\u969b\u5de5\u4f5c\u6eff\u8db3\u7387": "work_fulfillment_rate",
    "\u5b9e\u9645\u5de5\u4f5c\u6ee1\u8db3\u7387": "work_fulfillment_rate",
}

# 排除包含这些关键字的行（避免把平均行当指标行）
EXCLUDE_KEYWORDS = ["\u5e73\u5747", "\u5f53\u5468", "\u63a7\u5236", "\u7bc0\u7701/"]


class ParseReport:
    def __init__(self):
        self.iso_week: str = ""
        self.start_date: date = None
        self.end_date: date = None
        self.warehouses_found: list[str] = []
        self.new_warehouses: list[str] = []
        self.missing_warehouses: list[str] = []
        self.records_parsed: int = 0


def _match_metric(val_str: str) -> str | None:
    s = val_str.strip()
    # 排除平均行等
    for kw in EXCLUDE_KEYWORDS:
        if kw in s:
            return None
    if s in METRIC_MAP:
        return METRIC_MAP[s]
    for key, field in METRIC_MAP.items():
        if s.startswith(key) or key in s:
            return field
    return None


async def parse_excel(file_path: str, db: AsyncSession, uploaded_by: int, force_overwrite: bool = False) -> ParseReport:
    wb = load_workbook(file_path, data_only=True)
    sheet = None
    for ws in wb.worksheets:
        if "\u5173\u952e\u6570\u636e\u6c47\u603b" in ws.title or "\u95dc\u9375\u6578\u64da\u532f\u7e3d" in ws.title or "\u5173\u952e\u6570\u636e" in ws.title:
            sheet = ws
            break
    if not sheet:
        raise AppException(400, "\u627e\u4e0d\u5230\u300c\u5173\u952e\u6570\u636e\u6c47\u603b\u300dSheet")

    report = ParseReport()

    # \u626b\u63cf\u524d5\u884c\u8bc6\u522b\u4ed3\u5e93\u4ee3\u7801\u5217
    warehouse_cols: dict[int, str] = {}
    for row in sheet.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if WAREHOUSE_CODE_PATTERN.match(val):
                warehouse_cols[cell.column] = val
                if val not in report.warehouses_found:
                    report.warehouses_found.append(val)

    if not warehouse_cols:
        raise AppException(400, "\u672a\u8bc6\u522b\u5230\u4efb\u4f55\u4ed3\u5e93\u4ee3\u7801")

    result = await db.execute(select(Warehouse))
    existing_warehouses = {w.code: w for w in result.scalars().all()}
    db_codes = set(existing_warehouses.keys())
    excel_codes = set(warehouse_cols.values())

    for code in excel_codes - db_codes:
        wh = Warehouse(code=code, name=code)
        db.add(wh)
        await db.flush()
        existing_warehouses[code] = wh
        report.new_warehouses.append(code)

    report.missing_warehouses = list(db_codes - excel_codes)

    # \u626b\u63cf\u6307\u6807\u884c\uff1a\u6307\u6807\u540d\u5728 A \u5217
    metric_rows: list[tuple[str, int]] = []
    for row in sheet.iter_rows(min_row=1, values_only=False):
        a_cell = row[0] if row else None
        if a_cell and a_cell.value:
            val = str(a_cell.value).strip()
            matched = _match_metric(val)
            if matched:
                metric_rows.append((matched, a_cell.row))

    if not metric_rows:
        raise AppException(400, "\u672a\u8bc6\u522b\u5230\u4efb\u4f55\u6307\u6807\u884c")

    all_dates: set[date] = set()
    parsed_records: dict[tuple[date, str], dict] = {}

    for metric_field, row_idx in metric_rows:
        # \u4ece offset=0 \u5f00\u59cb\uff0c\u56e0\u4e3a\u6307\u6807\u884c\u672c\u8eab\u5c31\u5305\u542b\u7b2c\u4e00\u5929\u7684\u6570\u636e
        # \u65e5\u671f\u5728 B \u5217\uff0c\u6307\u6807\u540d\u5728 A \u5217
        for offset in range(0, 8):
            if row_idx + offset > sheet.max_row:
                break
            data_row = sheet[row_idx + offset]
            # \u65e5\u671f\u5728 B \u5217
            date_cell = data_row[1] if len(data_row) > 1 else None
            if not date_cell or not date_cell.value:
                continue

            row_date = _parse_date(date_cell.value)
            if not row_date:
                continue
            if row_date.weekday() == 6:  # \u8df3\u8fc7\u5468\u65e5
                continue

            all_dates.add(row_date)
            dow = row_date.weekday() + 1

            for col_idx, wh_code in warehouse_cols.items():
                cell_val = sheet.cell(row=row_idx + offset, column=col_idx).value
                if cell_val is None:
                    continue
                try:
                    val = float(cell_val)
                except (ValueError, TypeError):
                    continue

                key = (row_date, wh_code)
                if key not in parsed_records:
                    parsed_records[key] = {"date": row_date, "day_of_week": dow, "warehouse_code": wh_code}
                parsed_records[key][metric_field] = val

    if not all_dates:
        raise AppException(400, "\u672a\u89e3\u6790\u5230\u4efb\u4f55\u65e5\u671f\u6570\u636e")

    report.start_date = min(all_dates)
    report.end_date = max(all_dates)
    iso_calendar = report.start_date.isocalendar()
    report.iso_week = f"{iso_calendar[0]}-W{iso_calendar[1]:02d}"

    existing_report = await db.execute(
        select(WeeklyReport).where(WeeklyReport.iso_week == report.iso_week)
    )
    existing = existing_report.scalar_one_or_none()
    if existing and not force_overwrite:
        raise AppException(409, f"Week {report.iso_week} already exists", {"conflict": True, "iso_week": report.iso_week})

    if existing and force_overwrite:
        await db.execute(delete(DailyRecord).where(DailyRecord.weekly_report_id == existing.id))
        await db.execute(delete(WeeklyReport).where(WeeklyReport.id == existing.id))
        await db.flush()

    weekly_report = WeeklyReport(
        iso_week=report.iso_week,
        start_date=report.start_date,
        end_date=report.end_date,
        filename=Path(file_path).name,
        uploaded_by=uploaded_by,
    )
    db.add(weekly_report)
    await db.flush()

    for (row_date, wh_code), data in parsed_records.items():
        wh = existing_warehouses[wh_code]
        record = DailyRecord(
            weekly_report_id=weekly_report.id,
            warehouse_id=wh.id,
            date=data["date"],
            day_of_week=data["day_of_week"],
            iso_week=report.iso_week,
            system_headcount=data.get("system_headcount"),
            actual_attendance=data.get("actual_attendance"),
            required_headcount_so=data.get("required_headcount_so"),
            labor_savings=data.get("labor_savings"),
            work_fulfillment_rate=data.get("work_fulfillment_rate"),
        )
        db.add(record)
        report.records_parsed += 1

    await db.commit()
    return report


def _parse_date(val) -> date | None:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ["%m/%d/%Y", "%Y/%m/%d", "%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"]:
        try:
            d = datetime.strptime(s, fmt).date()
            if fmt == "%m/%d":
                d = d.replace(year=datetime.now().year)
            return d
        except ValueError:
            continue
    return None
